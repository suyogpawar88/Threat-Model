#!/usr/bin/env python3
"""
Excel (.xlsx) risk register builder for appsec-threat-modeler.

Takes the same structured JSON payload as build_docx_report.py and renders
a multi-sheet workbook: Summary, Threat Register, Risk Prioritization,
Compensating Controls, Attack Chains, Compliance Gaps, Data Flows,
Recommendations, Sources. Designed for risk owners who want to filter/sort/
pivot rather than read a narrative document.

Usage:
  python3 build_xlsx_report.py report_data.json Output_Risk_Register.xlsx
"""
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
RISK_FILL = {
    "High": PatternFill(start_color="F8CECC", end_color="F8CECC", fill_type="solid"),
    "Medium": PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid"),
    "Low": PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid"),
}
THIN_BORDER = Border(*(Side(style="thin", color="CCCCCC"),) * 4)
WRAP = Alignment(wrap_text=True, vertical="top")


def _write_sheet(ws, headers, rows, widths=None, risk_col=None):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        ws.append(row)
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = WRAP
            if risk_col is not None and c == risk_col + 1:
                val = str(cell.value or "")
                for label, fill in RISK_FILL.items():
                    if label.lower() in val.lower():
                        cell.fill = fill
                        break
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def build_workbook(data: dict, output_path: str):
    wb = Workbook()

    # --- Summary ---
    ws = wb.active
    ws.title = "Summary"
    threats = data.get("threats", [])
    counts = {b: sum(1 for t in threats if t.get("risk_bucket") == b) for b in ("High", "Medium", "Low")}
    rows = [
        ["Service", data.get("service_name", "")],
        ["Date", data.get("date", "")],
        ["Status", data.get("status", "")],
        ["Methodology", data.get("methodology", "")],
        ["Total Threats", len(threats)],
        ["High Risk", counts["High"]],
        ["Medium Risk", counts["Medium"]],
        ["Low Risk", counts["Low"]],
        ["Attack Chains Mapped", len(data.get("attack_chains", []))],
        ["Compliance Gaps Identified", len(data.get("compliance_gaps", []))],
    ]
    ws.append(["Field", "Value"])
    for c in (1, 2):
        ws.cell(row=1, column=c).fill = HEADER_FILL
        ws.cell(row=1, column=c).font = HEADER_FONT
    for row in rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50

    # --- Threat Register ---
    ws = wb.create_sheet("Threat Register")
    _write_sheet(
        ws,
        ["ID", "Category", "Title", "Description", "Affected Flow", "Threat Actor", "Likelihood", "Impact", "Score", "Risk", "Existing Controls", "Mitigation"],
        [
            [t.get("id"), t.get("category"), t.get("title"), t.get("description"), t.get("affected_flow"), t.get("threat_actor"), t.get("likelihood"), t.get("impact"), t.get("score"), t.get("risk_bucket"), t.get("existing_controls"), t.get("mitigation")]
            for t in threats
        ],
        widths=[6, 14, 20, 40, 12, 16, 9, 8, 7, 8, 30, 35],
        risk_col=9,
    )

    # --- Risk Prioritization ---
    ws = wb.create_sheet("Risk Prioritization")
    sorted_threats = sorted(threats, key=lambda t: (t.get("score", 0), t.get("likelihood", 0)), reverse=True)
    _write_sheet(
        ws,
        ["Rank", "ID", "Title", "Score", "Risk", "Rationale", "Owner Action"],
        [[i + 1, t.get("id"), t.get("title"), t.get("score"), t.get("risk_bucket"), t.get("impact_rationale"), t.get("owner_action")] for i, t in enumerate(sorted_threats)],
        widths=[6, 6, 22, 7, 8, 40, 30],
        risk_col=4,
    )

    # --- Compensating Controls ---
    ws = wb.create_sheet("Compensating Controls")
    _write_sheet(
        ws,
        ["ID", "Existing Compensating Controls", "Probability of Successful Mitigation", "Residual Risk", "Additional Controls Required"],
        [
            [
                t.get("id"),
                t.get("existing_controls", "None identified"),
                t.get("probability_of_mitigation", ""),
                t.get("residual_risk", ""),
                "; ".join(t.get("additional_controls_required", [])) if isinstance(t.get("additional_controls_required"), list) else t.get("additional_controls_required", ""),
            ]
            for t in threats
        ],
        widths=[6, 35, 22, 14, 40],
    )

    # --- Attack Chains ---
    ws = wb.create_sheet("Attack Chains")
    chain_rows = []
    for chain in data.get("attack_chains", []):
        for step in chain.get("steps", []):
            chain_rows.append([chain.get("chain_id"), chain.get("name"), step.get("step_num"), step.get("threat_id"), step.get("description"), chain.get("overall_risk")])
    _write_sheet(
        ws,
        ["Chain ID", "Chain Name", "Step", "Threat ID", "Step Description", "Overall Chain Risk"],
        chain_rows,
        widths=[10, 25, 6, 10, 45, 16],
        risk_col=5,
    )

    # --- Compliance Gaps ---
    ws = wb.create_sheet("Compliance Gaps")
    _write_sheet(
        ws,
        ["Control", "Compliance Status", "Gap Description", "Recommended Additional Controls"],
        [
            [
                g.get("control_name"),
                g.get("compliance_status"),
                g.get("gap_description"),
                "; ".join(g.get("recommended_additional_controls", [])) if isinstance(g.get("recommended_additional_controls"), list) else g.get("recommended_additional_controls", ""),
            ]
            for g in data.get("compliance_gaps", [])
        ],
        widths=[22, 18, 40, 40],
    )

    # --- Data Flows ---
    ws = wb.create_sheet("Data Flows")
    _write_sheet(
        ws,
        ["#", "Source", "Destination", "Data / Action", "Protocol", "Threats"],
        [[f.get("num"), f.get("source"), f.get("destination"), f.get("data_action"), f.get("protocol"), ", ".join(f.get("threats", []))] for f in data.get("data_flows", [])],
        widths=[4, 18, 18, 30, 14, 18],
    )

    # --- Recommendations ---
    ws = wb.create_sheet("Recommendations")
    _write_sheet(ws, ["#", "Recommendation"], [[i + 1, r] for i, r in enumerate(data.get("recommendations", []))], widths=[4, 90])

    # --- Sources ---
    ws = wb.create_sheet("Sources")
    src_rows = []
    for system, items in data.get("sources", {}).items():
        for item in items:
            src_rows.append([system.title(), item.get("ref", ""), item.get("description", "")])
    _write_sheet(ws, ["System", "Reference", "Description"], src_rows, widths=[14, 20, 60])

    wb.save(output_path)
    return output_path


def main():
    if len(sys.argv) != 3:
        print("Usage: python3 build_xlsx_report.py <report_data.json> <output.xlsx>", file=sys.stderr)
        sys.exit(1)
    with open(sys.argv[1], "r", encoding="utf-8") as f:
        data = json.load(f)
    path = build_workbook(data, sys.argv[2])
    print(f"Wrote {path}")


if __name__ == "__main__":
    main()
